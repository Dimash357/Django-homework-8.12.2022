from django.contrib.auth import logout, authenticate
from django.contrib.auth.hashers import make_password
from django.contrib.auth.models import User, Group
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.urls import reverse
from django.views import View
from django_app import models
from .forms import ProfileUpdateForm
import os
import sqlite3
import openpyxl



# Create your views here.


class HomeView(View):  # TODO контроллер класс
    template_name = 'django_app/home.html'

    def get(self, request: HttpRequest) -> HttpResponse:
        context = {}
        # return HttpResponse(content=b"<h1>Hello World</h1>")
        # return JsonResponse(data={"response": 'res'}, safe=True)
        return render(request, 'django_app/home.html', context=context)

    def post(self, request: HttpRequest) -> HttpResponse:
        context = {}
        # return HttpResponse(content=b"<h1>Hello World</h1>")
        # return JsonResponse(data={"response": 'res'}, safe=True)
        return render(request, 'django_app/home.html', context=context)


def home_view(request: HttpRequest) -> HttpResponse:  # TODO контроллер функция
    context = {}
    # return HttpResponse(content=b"<h1>Hello World</h1>")
    # return JsonResponse(data={"response": 'res'}, safe=True)
    return render(request, 'django_app/home.html', context=context)


def register(request: HttpRequest) -> HttpResponse:
    #

    if request.method == "GET":
        context = {}
        return render(request, 'django_app/register.html', context=context)
    elif request.method == "POST":

        # TODO получить с формы данные
        first_name = request.POST.get('first_name', "")
        last_name = request.POST.get('last_name', "")
        username = request.POST.get('username', None)
        password1 = request.POST.get('password1', "")
        password2 = request.POST.get('password2', "")

        if password1 and password1 != password2:
            raise Exception("пароли не совпадают!")
        if username and password1:
            User.objects.create(
                first_name=first_name,
                last_name=last_name,
                username=username,
                password=make_password(password1),
            )
            return redirect(reverse('django_app:login', args=()))
        else:
            raise Exception("данные не заполнены!")


def login(request: HttpRequest) -> HttpResponse:
    if request.method == "POST":

        email = request.POST.get("email")
        password = request.POST.get("password")

        print(f"email: {email}")
        print(f"password: {password}")

        with open('static/temp/data.txt', 'w') as file:
            file.write(f"{email}\n")
            file.write(f"{password}\n")
    context = {}
    return render(request, 'django_app/login.html', context=context)


def post_list(request: HttpRequest) -> HttpResponse:
    posts = models.Post.objects.all()  # filter order_by
    context = {"posts": posts}
    return render(request, 'django_app/post_list.html', context=context)


def post_detail(request: HttpRequest, pk: int) -> HttpResponse:
    post = models.Post.objects.get(id=pk)
    context = {"post": post}
    return render(request, 'django_app/post_detail.html', context=context)


def post_delete(request: HttpRequest, pk: int) -> HttpResponse:
    post = models.Post.objects.get(id=pk)
    post.delete()
    return redirect(reverse('django_app:post_list', args=()))


def post_pk_view(request: HttpRequest, pk: int) -> HttpResponse:
    if request.method == "GET":
        # post_list = models.Post.objects.all()
        # print(f"post_list: {post_list}")
        # context = {"post_list": post_list}
        context = {}
        return render(request, 'django_app/post_detail.html', context=context)
    context = {}
    # return HttpResponse(content=b"<h1>Hello World</h1>")
    # return JsonResponse(data={"response": 'res'}, safe=True)
    return render(request, 'django_app/post_list.html', context=context)


def home_main(request: HttpRequest) -> HttpResponse:
    context = {}
    return render(request, 'django_app/home_main.html', context=context)


def post_comment_create(request: HttpRequest, pk: int) -> HttpResponse:
    if request.method == "POST":
        text = request.POST.get('text', None)
        post = models.Post.objects.get(id=pk)  # определить, к какой статье создали комментарий
        models.PostComment.objects.create(
            user=request.User,
            article=post,
            text=text,
            # date_time=timezone.now(), # у нас стоит default
        )
        return redirect(reverse('django_app:post_detail', args=(pk,)))


def post_comment_delete(request: HttpRequest, pk: int) -> HttpResponse:
    comment = models.PostComment.objects.get(id=pk)
    pk = comment.article.id
    comment.delete()
    return redirect(reverse('django_app:post_detail', args=(pk,)))


def post_create(request: HttpRequest) -> HttpResponse:
    if request.method == "GET":
        context = {}
        return render(request, 'django_app/post_create.html', context=context)
    elif request.method == "POST":
        print("request: ", request)
        # print("request.data: ", request.data)
        print("request.POST: ", request.POST)
        print("request.GET: ", request.GET)
        print("request.META: ", request.META)

        title = request.POST.get('title', None)
        description = request.POST.get('description', "")
        post = models.Post.objects.create(
            title=title,
            description=description,
        )
        return redirect(reverse('django_app:post_list', args=()))


def logout_f(request: HttpRequest) -> HttpResponse:
    logout(request)
    return redirect(reverse('django_app:login', args=()))


def profile(request):
    return render(request, 'django_app/profile.html')


def profileupdate(request):
    if request.method == 'POST':
        pform = ProfileUpdateForm(request.POST, request.FILES, instance=request.user.profile)
        if pform.is_valid:
            pform.save()
            return render(request, 'django_app/profile.html')
    else:
        pform = ProfileUpdateForm(instance=request.user.profile)
    return render(request, 'django_app/profileupdate.html', {'pform': pform})


def json_page(request):
    users = [{"name": f"Dimash ({x})", "age": x} for x in range(1, 8)]
    print(users)

    return JsonResponse({"Your information": users})


def todo_create(request: HttpRequest) -> HttpResponse:
    if request.method == "GET":
        context = {}
        return render(request, 'django_app/todo_create.html', context=context)
    elif request.method == "POST":
        print("request: ", request)
        print("request.POST: ", request.POST)
        print("request.GET: ", request.GET)
        print("request.META: ", request.META)

        title = request.POST.get('title', "")
        description = request.POST.get('description', "")
        todo = models.Todo.objects.create(
            title=title,
            description=description,
        )
        return redirect(reverse('django_app:todo_list', args=()))

class CustomPaginator:
    @staticmethod
    def paginate(object_list: any, per_page=5, page_number=1):
        paginator_instance = Paginator(object_list=object_list, per_page=per_page)
        try:
            page = paginator_instance.page(number=page_number)
        except PageNotAnInteger:
            page = paginator_instance.page(number=1)
        except EmptyPage:
            page = paginator_instance.page(number=paginator_instance.num_pages)
        return page


def todo_list(request: HttpRequest) -> HttpResponse:
    todos = models.Todo.objects.all()

    selected_page_number = request.GET.get('page', 1)
    selected_limit_objects_per_page = request.GET.get('limit', 3)
    if request.method == "POST":
        selected_page_number = 1
        selected_limit_objects_per_page = 9999

        search_by_title = request.POST.get('search', None)
        if search_by_title is not None:
            todos = todos.filter(title__contains=str(search_by_title))

    page = CustomPaginator.paginate(
        object_list=todos, per_page=selected_limit_objects_per_page, page_number=selected_page_number
    )

    context = {"page": page}
    return render(request, 'django_app/todo_list.html', context=context)


def todo_delete(request: HttpRequest, pk: int) -> HttpResponse:
    todo = models.Todo.objects.get(id=pk)
    todo.delete()
    return redirect(reverse('django_app:todo_list', args=()))


def controller_test(request: HttpRequest) -> HttpResponse:
    context = {}
    return render(request, 'django_app/testcontroller.html', context=context)


def all_users(request):
    users = User.objects.all()

    selected_page_number = request.GET.get('page', 1)
    selected_limit_objects_per_page = request.GET.get('limit', 3)
    if request.method == "POST":
        selected_page_number = 1
        selected_limit_objects_per_page = 9999

        search_by_user = request.POST.get('search', None)
        if search_by_user is not None:
            users = users.filter(username__contains=str(search_by_user))

        # TODO просто расскаментируйте внизу и фильтр будет работать, не знаю в чем проблема,
        #  но они работают только по отдельности

        # filter_by_user = request.POST.get('filter', None)
        # if filter_by_user is not None:
        #     users = users.filter(username=filter_by_user)

    page = CustomPaginator.paginate(
        object_list=users, per_page=selected_limit_objects_per_page, page_number=selected_page_number
    )

    context = {"page": page, "users": users}
    return render(request, 'django_app/all_users.html', context=context)


def database(request):
    context = {}
    return render(request, 'django_app/database.html', context=context)


def task_delete(request: HttpRequest, pk: int) -> HttpResponse:
    task = models.Task.objects.get(id=pk)
    task.delete()
    return redirect(reverse('django_app:database', args=()))

# def todo_list_pygt6(request):
#     qt_creator_file = "window1.ui"
#     Ui_MainWindow, QtBaseClass = uic.loadUiType(qt_creator_file)
#
#
#     class TodoModel(QtCore.QAbstractListModel):
#         def __init__(self, *args, todos=None, **kwargs):
#             super(TodoModel, self).__init__(*args, **kwargs)
#             self.todos = todos or []
#
#         def data(self, index, role):
#             if role == Qt.DisplayRole:
#                 _, text = self.todos[index.row()]
#                 return text
#
#         def rowCount(self, index):
#             return len(self.todos)
#
#
#     class Window1(QtWidgets.QMainWindow, Ui_MainWindow):
#         def __init__(self):
#             QtWidgets.QMainWindow.__init__(self)
#             Ui_MainWindow.__init__(self)
#             self.setupUi(self)
#             self.model = TodoModel()
#             self.load()
#             self.todoView.setModel(self.model)
#             self.addButton.pressed.connect(self.add)
#             self.deleteButton.pressed.connect(self.delete)
#             self.completeButton.pressed.connect(self.complete)
#
#         def add(self):
#             text = self.todoEdit.text()
#             if text:
#                 self.model.todos.append((False, text))
#                 self.model.layoutChanged.emit()
#                 self.todoEdit.setText("")
#                 self.save()
#
#         def delete(self):
#             indexes = self.todoView.selectedIndexes()
#             if indexes:
#                 index = indexes[0]
#                 del self.model.todos[index.row()]
#                 self.model.layoutChanged.emit()
#                 self.todoView.clearSelection()
#                 self.save()
#
#         def complete(self):
#             indexes = self.todoView.selectedIndexes()
#             if indexes:
#                 index = indexes[0]
#                 row = index.row()
#                 status, text = self.model.todos[row]
#                 self.model.todos[row] = (True, text)
#                 self.model.dataChanged.emit(index, index)
#                 self.todoView.clearSelection()
#                 self.save()
#
#         def load(self):
#             try:
#                 with open('data.db', 'r') as f:
#                     self.model.todos = json.load(f)
#             except Exception:
#                 pass
#
#         def save(self):
#             with open('data.db', 'w') as f:
#                 data = json.dump(self.model.todos, f)
#
#
#     app = QtWidgets.QApplication(sys.argv)
#     window = Window1()
#     window.show()
#     app.exec_()
#     return redirect(reverse('django_app:todo_list_pygt6', args=()))


def hello(request: HttpRequest) -> HttpResponse:

    data_arr = tuple(({"id": x, "name": f"{x}", "amount": x**2} for x in range(1, 3)))
    print(data_arr, type(data_arr))

    context = {"data_arr": data_arr}

    return render(request, "django_app/hello.html", context)


def create(request: HttpRequest) -> HttpResponse:

    label = str(request.POST.get("name", ""))
    amount = int(request.POST.get("amount", 0))
    not_bubble_price = int(request.POST.get("first_cost", 0))
    bubble_percentage = int(request.POST.get("percent", 0))

    final_price = not_bubble_price + (round(not_bubble_price * 12 / 100, 0))
    vat_price = not_bubble_price + (round(not_bubble_price * 12 / 100, 0))
    overall = final_price * amount

    models.Product.objects.create(
        label=label,
        amount=amount,
        not_bubble_price=not_bubble_price,
        bubble_percentage=bubble_percentage,
        final_price=final_price,
        vat_price=vat_price,
        overall=overall,
    )
    return redirect(reverse('django_app:hello', args=()))


def delete(request: HttpRequest, pk: int) -> HttpResponse:
    one_todo = models.Product.objects.get(id=pk)
    one_todo.delete()
    return redirect(reverse('django_app:hello', args=()))


def import_data(request: HttpRequest) -> HttpResponse:
    excel_file = request.FILES["excel_file"]
    if excel_file:
        workbook = openpyxl.load_workbook(excel_file)
        worksheet = workbook.active  # sheet = file_to_read["Лист1", "Sheet2", "Sheet3"]

        matrix = []
        for row in range(2, worksheet.max_row + 1):
            data = []
            for column in range(1, worksheet.max_column + 1):
                value = worksheet.cell(row, column).value

                if value is None or value == "-" or value == "":
                    value = 0

                if column == 1:
                    value = str(value)
                else:
                    value = int(str(value).strip())

                data.append(value)
            matrix.append(data)

        print(matrix)
        for product in matrix:

            label = str(product[0])
            amount = int(product[1])
            not_bubble_price = int(product[2])
            bubble_percentage = int(product[3])

            final_price = not_bubble_price + (round(not_bubble_price * 12 / 100, 0))
            vat_price = not_bubble_price + (round(not_bubble_price * 12 / 100, 0))
            overall = final_price * amount

            models.Product.objects.create(
                label=label,
                amount=amount,
                not_bubble_price=not_bubble_price,
                bubble_percentage=bubble_percentage,
                final_price=final_price,
                vat_price=vat_price,
                overall=overall,
            )

    # TODO запись в базу данных:
    # SQL

    return redirect(reverse('django_app:hello', args=()))

    # prj_dir = os.path.dirname(os.path.dirname(os.path.abspath(file)))  # Получение ДБ, не уверен

    # # 1.Сonnecting to the database

    # # file

    # connect = sqlite3.connect(prj_dir + "/")

    # cursor = connect.cursor()

    # # 2. Working with xlsx file

    # for row in range(2, sheet.max_row + 1):

    #     data = []

    #     for col in range(1, 5):
    #         value = sheet.cell(row, col).value

    #         data.append(value)

    #     # 3. Writing to the database and closing the connection

    #     cursor.execute("INSERT INTO ... VALUES (?, ?, ?, ?);", (data[0], data[1], data[2], data[3]))

    # # Save change
    # connect.commit()
    # # Close connection
    # connect.close()


def export_data(request: HttpRequest) -> HttpResponse:
    products = models.Product.objects.all()
    print(products, type(products))

    matrix = []

    for row_index, product in enumerate(products, 1):
        label = product.label
        amount = product.amount
        not_bubble_price = product.not_bubble_price
        bubble_percentage = product.bubble_percentage
        final_price = product.final_price
        vat_price = product.vat_price
        overall = product.overall

        row = [label, amount, not_bubble_price, bubble_percentage, final_price, vat_price, overall]
        matrix.append(row)
        print(matrix)
        if overall > 999:
            print(f"{label}: [{amount}]  = ", '{0:,}'.format(overall).replace(',', ' '))
        elif overall < 999:
            print(f"{label}: [{amount}]  = {overall}")

    return redirect(reverse("django_app:hello", args=()))

